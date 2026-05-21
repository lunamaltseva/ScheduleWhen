#!/usr/bin/env python3
"""
Generate the Fall 2026 hypothetical-student enrollment table (1,222 students).

Rules implemented:
  * 1,222 students: 84% undergraduate (1,028), 16% masters (194).
  * 450 international students spread across all years/programs.
      - international -> RFL (Russian as Foreign) + KLL-104 (Kyrgyz for foreigners)
      - domestic      -> RUS (Russian) + KLL-203/103 (Kyrgyz)
  * Credit loads: 24-33 credits; most students reach the cap.
      - Freshman cap = 30 credits; Sophomore/Junior/Senior cap = 33.
      - Masters fill their program load (typically 18-30).
  * Every course is written with its section and meeting time.
  * No two of a student's courses overlap in time (conflict-free schedules).
  * Section seat counts are treated as soft (an "add form" allows a few extra,
    and sections are often under-filled) so capacity is not hard-enforced.
"""

import sys
import os
sys.path.insert(0, '/home/lunamaltseva/.local/lib/python3.14/site-packages')

import random
from collections import Counter

_ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import scripts.course_data as CD
import scripts.dept_data as DD

random.seed(42)

# ===========================================================================
# Build a global course catalog: code -> course dict (keep richest version)
# ===========================================================================

CATALOG = {}

def _register(courses):
    for c in courses:
        prev = CATALOG.get(c["code"])
        if prev is None or len(c["sections"]) > len(prev["sections"]):
            CATALOG[c["code"]] = c

_register(CD.GENED)
_register(CD.SPORTS)
_register(CD.RUS_DOMESTIC)
_register(CD.RFL_INTERNATIONAL)
_register(CD.KLL_DOMESTIC)
_register(CD.KLL_INTERNATIONAL)
for _dept_courses in DD.DEPT_CATALOG.values():
    _register(_dept_courses)

GENED_CODES = [c["code"] for c in CD.GENED]

# Tracks how many students have enrolled in each (course_code, section_label).
ENROLLMENT = Counter()

# ===========================================================================
# Helpers
# ===========================================================================

def fmt_slots(slots):
    """Render meeting blocks like 'Mon/Wed 9:25' or 'Fri 9:25 & 10:50'."""
    if not slots:
        return "Individual"
    times = {t for _, t in slots}
    if len(times) == 1:
        days = [d for d, _ in slots]
        return f"{'/'.join(days)} {next(iter(times))}"
    return " & ".join(f"{d} {t}" for d, t in slots)

def course_label(course, section):
    return (f"{course['code']} (sec {section['sec']}): {course['name']} "
            f"— {fmt_slots(section['slots'])} — {course['credits']}cr")


class Schedule:
    """Accumulates a conflict-free set of courses for one student."""
    def __init__(self, cap):
        self.cap = cap
        self.occupied = set()      # set of (day, time)
        self.entries = []          # list of display strings
        self.credits = 0

    def _section_ok(self, course, section, ignore_seats=False):
        if self.credits + course["credits"] > self.cap:
            return False
        if not ignore_seats and ENROLLMENT[(course["code"], section["sec"])] >= section["seats"]:
            return False
        return all(s not in self.occupied for s in section["slots"])

    def add(self, code):
        course = CATALOG.get(code)
        if course is None:
            return False
        if any(e.startswith(code + " (") for e in self.entries):
            return False
        sections = list(course["sections"])
        random.shuffle(sections)
        # First pass: prefer sections that still have seats.
        for section in sections:
            if self._section_ok(course, section):
                self.occupied.update(section["slots"])
                self.entries.append(course_label(course, section))
                self.credits += course["credits"]
                ENROLLMENT[(code, section["sec"])] += 1
                return True
        # Second pass: overflow — add-form scenario, ignore seat cap.
        for section in sections:
            if self._section_ok(course, section, ignore_seats=True):
                self.occupied.update(section["slots"])
                self.entries.append(course_label(course, section))
                self.credits += course["credits"]
                ENROLLMENT[(code, section["sec"])] += 1
                return True
        return False


# ===========================================================================
# Department configuration  (cohort = students per year level)
#   UG x 4 years = 1,028 ; PG x 2 years = 194 ; total = 1,222
# ===========================================================================

UG_DEPTS = [
    ("BA",   None,   40), ("SFW",  None,   24), ("AMI",  None,   16),
    ("ECO",  None,   20), ("ICP",  None,   20),
    ("LAS",  "DC",    8), ("LAS",  "ES",    8), ("LAS",  "HR",    6),
    ("LAS",  "MC",   10), ("LAS",  "MM",    8), ("LAS",  "PC",   10),
    ("LAS",  "SEDT",  9), ("LAS",  "UPD",   9),
    ("JMC",  None,   10), ("PSY",  None,   13), ("SOC",  None,   10),
    ("ANTH", None,   10), ("ESCS", None,   10), ("TCMA", None,   10),
    ("IBL",  None,    6),
]   # 257 / year  ->  1,028

PG_DEPTS = [
    ("LLM",    None, 12), ("MAANTH", None, 15), ("MACAS",  None, 10),
    ("MAPAP",  None, 15), ("MAT",    None, 15), ("MBA",    None, 20),
    ("MSECO",  None, 10),
]   # 97 / year  ->  194

UG_YEARS = ["Freshman", "Sophomore", "Junior", "Senior"]
PG_YEARS = ["Masters 1st Year", "Masters 2nd Year"]

TOTAL = 1222
N_INTERNATIONAL = 450


def pick_target(cap):
    """Most students take the cap; a minority take lighter loads (>=24)."""
    r = random.random()
    if r < 0.62:
        return cap
    if r < 0.86:
        return cap - 3
    return random.choice([24, 27, min(30, cap)])


def russian_code(international):
    if international:
        return random.choice(["RFL-103.1", "RFL-203.1"])
    return random.choice(["RUS-101.1", "RUS-201.1"])

def kyrgyz_code(international):
    if international:
        return "KLL-104.1"
    return random.choice(["KLL-203.1", "KLL-103.1"])


def build_schedule(dept, conc, year, international):
    is_pg = year in PG_YEARS
    cap = 33
    if year == "Freshman":
        cap = 30
    target = pick_target(cap) if not is_pg else 33

    sch = Schedule(cap)

    # 1. Freshman writing cluster (FYS-100 4cr + ENG-122 6cr + FYS-100.3 2cr)
    if year == "Freshman":
        bundles = CD.fys1_bundle_sections()
        random.shuffle(bundles)
        fys_assigned = False
        for strict in (True, False):   # second pass = overflow / add-form
            for bundle in bundles:
                f_slots, e_slots = bundle["fys100_slots"], bundle["eng122_slots"]
                if strict and ENROLLMENT[("FYS-100", bundle["sec"])] >= bundle["seats"]:
                    continue
                if all(s not in sch.occupied for s in f_slots + e_slots) \
                   and sch.credits + 12 <= cap:
                    sch.occupied.update(f_slots + e_slots)
                    sch.entries.append(
                        f"FYS-100 (sec {bundle['sec']}): First Year Seminar I "
                        f"— {fmt_slots(f_slots)} — 4cr")
                    sch.entries.append(
                        f"ENG-122 (sec {bundle['sec']}): English Composition for "
                        f"Liberal Arts I — {fmt_slots(e_slots)} — 6cr")
                    sch.entries.append(
                        "FYS-100.3: First Year Seminar I: Philosophy "
                        "— Individual — 2cr")
                    ENROLLMENT[("FYS-100", bundle["sec"])] += 1
                    ENROLLMENT[("ENG-122", bundle["sec"])] += 1
                    sch.credits += 12
                    fys_assigned = True
                    break
            if fys_assigned:
                break

    # 2. Required major courses for this dept/concentration/year
    required = list(DD.REQUIRED.get((dept, conc, year), []))
    if dept == "LAS":
        required = DD.LAS_CORE_BY_YEAR.get(year, []) + required
    for code in required:
        sch.add(code)

    # 3. Sophomore language requirements (Kyrgyz 4cr + Russian 2cr)
    if year == "Sophomore":
        sch.add(kyrgyz_code(international))
        sch.add(russian_code(international))

    # 4. Fill to target with electives + GenEd (UG) / electives (PG)
    pool = list(DD.ELECTIVES.get(dept, []))
    if not is_pg:
        pool += GENED_CODES
    random.shuffle(pool)
    for code in pool:
        if sch.credits >= target:
            break
        sch.add(code)

    # PG: make sure at least the minimum full-time load (>=18) where possible
    if is_pg and sch.credits < 18:
        for code in DD.ELECTIVES.get(dept, []):
            if sch.credits >= 18:
                break
            sch.add(code)

    return sch


# ===========================================================================
# Generate students
# ===========================================================================

roster = []   # (dept, conc, year)
for dept, conc, cohort in UG_DEPTS:
    for year in UG_YEARS:
        roster += [(dept, conc, year)] * cohort
for dept, conc, cohort in PG_DEPTS:
    for year in PG_YEARS:
        roster += [(dept, conc, year)] * cohort

assert len(roster) == TOTAL, f"roster={len(roster)} (expected {TOTAL})"

# Flag 450 international students at random across the whole roster.
intl_flags = [True] * N_INTERNATIONAL + [False] * (TOTAL - N_INTERNATIONAL)
random.shuffle(intl_flags)

students = []
for i, ((dept, conc, year), international) in enumerate(zip(roster, intl_flags), 1):
    sch = build_schedule(dept, conc, year, international)
    students.append({
        "id": f"S{i:05d}",
        "dept": dept,
        "conc": conc or "",
        "year": year,
        "international": "Yes" if international else "No",
        "credits": sch.credits,
        "courses": sch.entries,
    })

# ===========================================================================
# Write Excel
# ===========================================================================

DEPT_COLORS = {
    "BA": "D6E4F0", "ECO": "D5E8D4", "AMI": "DAE8FC", "SFW": "E1D5E7",
    "ICP": "FFF2CC", "JMC": "FFE6CC", "PSY": "F8CECC", "SOC": "EADCE8",
    "ANTH": "FDE8C8", "ESCS": "C9EAD5", "TCMA": "FAD7D7", "IBL": "D7EAF8",
    "LAS": "FFFDE7", "LLM": "D8BFD8", "MAANTH": "B5D5C5", "MACAS": "B0C4DE",
    "MAPAP": "FFCCE5", "MAT": "FFFACD", "MBA": "C0D8F0", "MSECO": "C8E6C9",
}

MAX_COURSES = max(len(s["courses"]) for s in students)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Fall 2026 Students"

HDR_FILL = PatternFill("solid", fgColor="1F4E79")
HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
DATA_ALIGN = Alignment(horizontal="left", vertical="center")
thin = Side(style="thin", color="BDBDBD")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

headers = (["Student ID", "Department", "Concentration", "Year",
            "International", "Total Credits"]
           + [f"Course {i}" for i in range(1, MAX_COURSES + 1)])
widths = [11, 11, 13, 17, 12, 13] + [62] * MAX_COURSES

ws.row_dimensions[1].height = 30
for col, (h, w) in enumerate(zip(headers, widths), 1):
    cell = ws.cell(row=1, column=col, value=h)
    cell.fill, cell.font, cell.alignment, cell.border = HDR_FILL, HDR_FONT, HDR_ALIGN, BORDER
    ws.column_dimensions[get_column_letter(col)].width = w

for r, s in enumerate(students, 2):
    fill = PatternFill("solid", fgColor=DEPT_COLORS.get(s["dept"], "FFFFFF"))
    row = ([s["id"], s["dept"], s["conc"], s["year"], s["international"], s["credits"]]
           + s["courses"] + [""] * (MAX_COURSES - len(s["courses"])))
    for col, val in enumerate(row, 1):
        cell = ws.cell(row=r, column=col, value=val)
        cell.fill, cell.alignment, cell.border = fill, DATA_ALIGN, BORDER

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(students)+1}"

# Summary sheet
ws2 = wb.create_sheet("Summary")
for col, w in zip("ABCD", (20, 16, 12, 14)):
    ws2.column_dimensions[col].width = w
for col, h in enumerate(["Department", "Concentration", "Count", "Intl"], 1):
    cell = ws2.cell(row=1, column=col, value=h)
    cell.fill, cell.font, cell.alignment = HDR_FILL, HDR_FONT, HDR_ALIGN

counts = Counter((s["dept"], s["conc"]) for s in students)
intl_counts = Counter((s["dept"], s["conc"]) for s in students if s["international"] == "Yes")
for r, ((dept, conc), cnt) in enumerate(sorted(counts.items()), 2):
    ws2.cell(row=r, column=1, value=dept)
    ws2.cell(row=r, column=2, value=conc or "(main)")
    ws2.cell(row=r, column=3, value=cnt)
    ws2.cell(row=r, column=4, value=intl_counts.get((dept, conc), 0))
tr = len(counts) + 2
ws2.cell(row=tr, column=2, value="TOTAL").font = Font(bold=True)
ws2.cell(row=tr, column=3, value=len(students)).font = Font(bold=True)
ws2.cell(row=tr, column=4, value=sum(intl_counts.values())).font = Font(bold=True)

out = os.path.join(_ROOT, "public", "Fall_2026_Students.xlsx")
wb.save(out)

# ===========================================================================
# Console report + sanity checks
# ===========================================================================

ug = sum(1 for s in students if s["year"] in UG_YEARS)
pg = len(students) - ug
intl = sum(1 for s in students if s["international"] == "Yes")
cred = [s["credits"] for s in students]
ug_cred = [s["credits"] for s in students if s["year"] in UG_YEARS]
under = [s for s in students if s["credits"] < 24 and s["year"] not in PG_YEARS]
over_fresh = [s for s in students if s["year"] == "Freshman" and s["credits"] > 30]
over_any = [s for s in students if s["credits"] > 33]

print(f"Saved -> {out}")
print(f"  Total students   : {len(students)}")
print(f"  Undergraduate    : {ug} ({ug/len(students)*100:.1f}%)")
print(f"  Postgraduate     : {pg} ({pg/len(students)*100:.1f}%)")
print(f"  International     : {intl} ({intl/len(students)*100:.1f}%)")
print(f"  Credits min/avg/max (all): {min(cred)}/{sum(cred)/len(cred):.1f}/{max(cred)}")
print(f"  UG credit distribution   : "
      + ", ".join(f"{c}cr:{ug_cred.count(c)}" for c in sorted(set(ug_cred))))
print(f"  Max courses in a row     : {MAX_COURSES}")
print(f"  UG students under 24cr   : {len(under)}")
print(f"  Freshmen over 30cr       : {len(over_fresh)}")
print(f"  Anyone over 33cr         : {len(over_any)}")
